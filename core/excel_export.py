"""
SurplusIQ — Excel Exporter

Generates a professional multi-tab Excel workbook from the consolidated lead data.

Tabs:
  1. Summary       — Executive overview with county totals + score distribution
  2. All Leads     — All 124 qualifying leads, sorted by surplus
  3. A+ Priority   — Just the top-tier leads (≥$100K surplus)
  4. Florida       — Florida-only leads
  5. Ohio          — Ohio-only leads

Usage:
    python -m core.excel_export
    # → Outputs to data/output/SurplusIQ_Leads_<YYYY-MM-DD>.xlsx
"""

from __future__ import annotations
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from core.loader import load_all_leads, get_summary, COUNTY_INFO, PROJECT_ROOT


# ═══════════════════════════════════════════════════════════════════════
# Styling
# ═══════════════════════════════════════════════════════════════════════
HEADER_FILL    = PatternFill("solid", fgColor="1F4E78")  # deep navy blue
HEADER_FONT    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT     = Font(name="Calibri", size=18, bold=True, color="1F4E78")
SUBTITLE_FONT  = Font(name="Calibri", size=11, italic=True, color="555555")
LABEL_FONT     = Font(name="Calibri", size=10, bold=True, color="333333")
DATA_FONT      = Font(name="Calibri", size=10)
TOTAL_FONT     = Font(name="Calibri", size=11, bold=True, color="1F4E78")

ALT_FILL       = PatternFill("solid", fgColor="F4F7FB")  # light blue tint

SCORE_FILLS = {
    "A+": PatternFill("solid", fgColor="C6EFCE"),  # green
    "A":  PatternFill("solid", fgColor="DDEBF7"),  # light blue
    "B":  PatternFill("solid", fgColor="FFF2CC"),  # light yellow
    "C":  PatternFill("solid", fgColor="FCE4D6"),  # light orange
}
SCORE_FONTS = {
    "A+": Font(name="Calibri", size=10, bold=True, color="006100"),
    "A":  Font(name="Calibri", size=10, bold=True, color="1F4E78"),
    "B":  Font(name="Calibri", size=10, bold=True, color="9C5700"),
    "C":  Font(name="Calibri", size=10, bold=True, color="9C0006"),
}

THIN  = Side(border_style="thin", color="D0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")


# ═══════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════
def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _style_header_row(ws, row_num: int, n_cols: int):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row_num].height = 28


def _format_currency(cell):
    cell.number_format = '"$"#,##0'
    cell.alignment = RIGHT


# ═══════════════════════════════════════════════════════════════════════
# Tab 1: Executive Summary
# ═══════════════════════════════════════════════════════════════════════
def build_summary_tab(ws, leads, summary):
    ws.title = "Summary"

    # Title block
    ws["A1"] = "SurplusIQ — Surplus Funds Intelligence Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:G2")

    ws["A3"] = f"Coverage: 10 counties (5 Florida + 5 Ohio) | Daily auction surveillance"
    ws["A3"].font = SUBTITLE_FONT
    ws.merge_cells("A3:G3")

    # Top KPIs (row 5)
    ws["A5"] = "TOTAL LEADS"
    ws["A5"].font = LABEL_FONT
    ws["B5"] = summary["total_leads"]
    ws["B5"].font = TITLE_FONT

    ws["D5"] = "TOTAL SURPLUS"
    ws["D5"].font = LABEL_FONT
    ws["E5"] = summary["total_surplus"]
    ws["E5"].font = TITLE_FONT
    _format_currency(ws["E5"])

    ws["G5"] = f"A+ COUNT"
    ws["G5"].font = LABEL_FONT

    # By State (rows 7-10)
    ws["A7"] = "By State"
    ws["A7"].font = LABEL_FONT
    ws.merge_cells("A7:E7")
    ws["A7"].fill = HEADER_FILL
    ws["A7"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A7"].alignment = LEFT

    ws["A8"] = "State";    ws["B8"] = "Leads"; ws["C8"] = "Surplus"
    for col in ["A8", "B8", "C8"]:
        ws[col].font = LABEL_FONT
        ws[col].fill = ALT_FILL
        ws[col].border = BORDER

    row = 9
    for state, data in summary["by_state"].items():
        if data["leads"] > 0:
            ws.cell(row=row, column=1, value=state).font = DATA_FONT
            ws.cell(row=row, column=2, value=data["leads"]).font = DATA_FONT
            cell = ws.cell(row=row, column=3, value=data["surplus"])
            _format_currency(cell)
            cell.font = DATA_FONT
            row += 1

    # By County (starting row ~13)
    row += 2
    ws.cell(row=row, column=1, value="By County").font = LABEL_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    row += 1
    headers = ["County", "State", "Leads", "Total Surplus", "Top Lead"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = LABEL_FONT
        c.fill = ALT_FILL
        c.border = BORDER

    row += 1
    for c in summary["by_county"]:
        ws.cell(row=row, column=1, value=c["county_name"]).font = DATA_FONT
        ws.cell(row=row, column=2, value=c["state"]).font = DATA_FONT
        ws.cell(row=row, column=3, value=c["leads"]).font = DATA_FONT
        cell = ws.cell(row=row, column=4, value=c["surplus"]); _format_currency(cell); cell.font = DATA_FONT
        cell = ws.cell(row=row, column=5, value=c["top_lead"]); _format_currency(cell); cell.font = DATA_FONT
        row += 1

    # By Score
    row += 2
    ws.cell(row=row, column=1, value="By Lead Score").font = LABEL_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

    row += 1
    score_headers = ["Tier", "Count", "Threshold"]
    for i, h in enumerate(score_headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = LABEL_FONT
        c.fill = ALT_FILL
        c.border = BORDER

    row += 1
    score_thresholds = {"A+": "≥ $100,000", "A": "≥ $50,000", "B": "≥ $25,000", "C": "≥ $10,000"}
    for score in ["A+", "A", "B", "C"]:
        count = summary["by_score"].get(score, 0)
        cell_score = ws.cell(row=row, column=1, value=score)
        cell_score.fill = SCORE_FILLS[score]
        cell_score.font = SCORE_FONTS[score]
        cell_score.alignment = CENTER

        ws.cell(row=row, column=2, value=count).font = DATA_FONT
        ws.cell(row=row, column=3, value=score_thresholds[score]).font = DATA_FONT
        row += 1

    # Top 5 Leads
    row += 2
    ws.cell(row=row, column=1, value="Top 5 Leads").font = LABEL_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    row += 1
    top_headers = ["Rank", "Surplus", "Score", "County", "Case #", "Address"]
    for i, h in enumerate(top_headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = LABEL_FONT
        c.fill = ALT_FILL
        c.border = BORDER

    row += 1
    for i, l in enumerate(summary["top_5_leads"], 1):
        ws.cell(row=row, column=1, value=f"#{i}").font = DATA_FONT
        cell = ws.cell(row=row, column=2, value=l["surplus"]); _format_currency(cell); cell.font = DATA_FONT
        cell_score = ws.cell(row=row, column=3, value=l["score"])
        cell_score.fill = SCORE_FILLS.get(l["score"], PatternFill())
        cell_score.font = SCORE_FONTS.get(l["score"], DATA_FONT)
        cell_score.alignment = CENTER
        ws.cell(row=row, column=4, value=f"{l['county']}, {l['state']}").font = DATA_FONT
        ws.cell(row=row, column=5, value=l["case_number"]).font = DATA_FONT
        ws.cell(row=row, column=6, value=l["address"]).font = DATA_FONT
        row += 1

    _set_col_widths(ws, {"A": 18, "B": 14, "C": 8, "D": 22, "E": 22, "F": 50, "G": 16})


# ═══════════════════════════════════════════════════════════════════════
# Tab: Leads table (used for "All Leads", "A+ Priority", "Florida", "Ohio")
# ═══════════════════════════════════════════════════════════════════════
def build_leads_tab(ws, title: str, leads: list, subtitle: str = ""):
    ws.title = title

    headers = [
        "Rank", "Score", "Surplus", "Sale Price", "Opening Bid",
        "County", "State", "Case Number", "Sale Date",
        "Property Address", "Parcel ID", "Sold To", "Status",
    ]

    # Title row
    ws["A1"] = f"SurplusIQ — {title}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 32

    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    ws[f"A3"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | {len(leads)} leads"
    ws[f"A3"].font = SUBTITLE_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))

    # Header row at row 5
    HEADER_ROW = 5
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=HEADER_ROW, column=i, value=h)
    _style_header_row(ws, HEADER_ROW, len(headers))

    # Data rows
    for i, lead in enumerate(leads, 1):
        r = HEADER_ROW + i

        ws.cell(row=r, column=1, value=i).font = DATA_FONT
        ws.cell(row=r, column=1).alignment = CENTER

        # Score with color
        cell_score = ws.cell(row=r, column=2, value=lead.score)
        cell_score.fill = SCORE_FILLS.get(lead.score, PatternFill())
        cell_score.font = SCORE_FONTS.get(lead.score, DATA_FONT)
        cell_score.alignment = CENTER

        cell = ws.cell(row=r, column=3, value=lead.gross_surplus); _format_currency(cell); cell.font = Font(bold=True, size=10)
        cell = ws.cell(row=r, column=4, value=lead.final_sale_price); _format_currency(cell); cell.font = DATA_FONT
        cell = ws.cell(row=r, column=5, value=lead.opening_bid); _format_currency(cell); cell.font = DATA_FONT

        ws.cell(row=r, column=6,  value=lead.county_name).font = DATA_FONT
        ws.cell(row=r, column=7,  value=lead.state).font = DATA_FONT
        ws.cell(row=r, column=7).alignment = CENTER
        ws.cell(row=r, column=8,  value=lead.case_number).font = DATA_FONT
        ws.cell(row=r, column=9,  value=lead.sale_date).font = DATA_FONT
        ws.cell(row=r, column=10, value=lead.address).font = DATA_FONT
        ws.cell(row=r, column=11, value=lead.parcel_id).font = DATA_FONT
        ws.cell(row=r, column=12, value=lead.sold_to).font = DATA_FONT
        ws.cell(row=r, column=13, value=lead.auction_status).font = DATA_FONT

        # Alternating row fill
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                if not ws.cell(row=r, column=col).fill or ws.cell(row=r, column=col).fill.start_color.rgb in (None, "00000000"):
                    if col != 2:  # don't override score color
                        ws.cell(row=r, column=col).fill = ALT_FILL

        # Borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = BORDER

    # Total row at the bottom
    total_row = HEADER_ROW + len(leads) + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=2, value=f"{len(leads)} leads").font = TOTAL_FONT
    cell = ws.cell(row=total_row, column=3, value=sum(l.gross_surplus for l in leads))
    _format_currency(cell)
    cell.font = TOTAL_FONT
    cell = ws.cell(row=total_row, column=4, value=sum(l.final_sale_price for l in leads))
    _format_currency(cell)
    cell.font = TOTAL_FONT

    # Freeze panes (lock title + header)
    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    # Column widths
    widths = {
        "A": 6,    # rank
        "B": 7,    # score
        "C": 14,   # surplus
        "D": 14,   # sale price
        "E": 14,   # opening bid
        "F": 14,   # county
        "G": 6,    # state
        "H": 22,   # case number
        "I": 12,   # sale date
        "J": 40,   # address
        "K": 18,   # parcel
        "L": 18,   # sold to
        "M": 12,   # status
    }
    _set_col_widths(ws, widths)


# ═══════════════════════════════════════════════════════════════════════
# Main export
# ═══════════════════════════════════════════════════════════════════════
def export_excel(output_path: Path = None) -> Path:
    """
    Build the multi-tab Excel workbook.
    Returns the path of the saved file.
    """
    if output_path is None:
        out_dir = PROJECT_ROOT / "data" / "output"
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = out_dir / f"SurplusIQ_Leads_{date.today().isoformat()}.xlsx"

    print(f"📊 Loading leads...")
    leads = load_all_leads()
    summary = get_summary(leads)
    print(f"   ✓ {len(leads)} qualifying leads loaded")
    print(f"   ✓ ${summary['total_surplus']:,.0f} total surplus identified")

    print(f"📝 Building Excel workbook...")
    wb = Workbook()

    # Tab 1: Summary (default sheet)
    ws_summary = wb.active
    build_summary_tab(ws_summary, leads, summary)

    # Tab 2: All Leads
    ws_all = wb.create_sheet("All Leads")
    build_leads_tab(ws_all, "All Leads", leads,
                    subtitle="All qualifying leads (3rd-party wins ≥ $10K surplus), sorted by surplus")

    # Tab 3: A+ Priority
    aplus_leads = [l for l in leads if l.score == "A+"]
    if aplus_leads:
        ws_aplus = wb.create_sheet("A+ Priority")
        build_leads_tab(ws_aplus, "A+ Priority Leads", aplus_leads,
                        subtitle=f"Top-tier leads with surplus ≥ $100K — {len(aplus_leads)} leads totaling ${sum(l.gross_surplus for l in aplus_leads):,.0f}")

    # Tab 4: Florida
    fl_leads = [l for l in leads if l.state == "FL"]
    ws_fl = wb.create_sheet("Florida")
    build_leads_tab(ws_fl, "Florida Leads", fl_leads,
                    subtitle=f"Florida counties: Miami-Dade, Broward, Duval, Lee, Orange — {len(fl_leads)} leads totaling ${sum(l.gross_surplus for l in fl_leads):,.0f}")

    # Tab 5: Ohio
    oh_leads = [l for l in leads if l.state == "OH"]
    ws_oh = wb.create_sheet("Ohio")
    build_leads_tab(ws_oh, "Ohio Leads", oh_leads,
                    subtitle=f"Ohio counties: Cuyahoga, Franklin, Montgomery, Summit, Hamilton — {len(oh_leads)} leads totaling ${sum(l.gross_surplus for l in oh_leads):,.0f}")

    print(f"💾 Saving to {output_path}...")
    wb.save(output_path)
    print(f"   ✓ Saved successfully")
    print()
    print(f"📂 Open with: open '{output_path}'")
    return output_path


if __name__ == "__main__":
    path = export_excel()
    print()
    print("=" * 70)
    print(f"  ✓ Excel export complete")
    print(f"  📁 {path}")
    print("=" * 70)
