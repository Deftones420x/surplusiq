"""
SurplusIQ — Excel Export
Generates a formatted XLSX file from processed leads
"""

import json
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
except ImportError:
    raise ImportError("Run: pip install openpyxl")

ROOT       = Path(__file__).parent.parent
DATA_DIR   = ROOT / "data"
OUTPUT_DIR = ROOT / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Color Palette ─────────────────────────────────────────────────────
CLR_BG_DARK   = "0D1117"
CLR_GREEN     = "00E5A0"
CLR_BLUE      = "38BEFF"
CLR_AMBER     = "FFB830"
CLR_RED       = "FF5370"
CLR_WHITE     = "FFFFFF"
CLR_GRAY_LIGHT = "F5F7FA"
CLR_HEADER_BG = "1B4F8A"
CLR_GRADE_AP  = "D4EFDF"
CLR_GRADE_A   = "EBF5FB"
CLR_GRADE_B   = "FEF9E7"
CLR_GRADE_C   = "FDEDEC"
CLR_ROW_ALT   = "F8FAFB"

def grade_fill(grade: str) -> PatternFill:
    colors = {
        "A+": CLR_GRADE_AP,
        "A":  CLR_GRADE_A,
        "B":  CLR_GRADE_B,
        "C":  CLR_GRADE_C,
    }
    hex_color = colors.get(grade, CLR_WHITE)
    return PatternFill("solid", fgColor=hex_color)

def claim_fill(status: str) -> PatternFill:
    colors = {
        "none":      "D4EFDF",
        "partial":   "FEF9E7",
        "filed":     "FDEBD0",
        "disbursed": "FADBD8",
        "n/a":       "F2F3F4",
    }
    return PatternFill("solid", fgColor=colors.get(status, CLR_WHITE))

def thin_border() -> Border:
    thin = Side(style="thin", color="DDDDDD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=CLR_HEADER_BG)


COLUMNS = [
    # (header, field_key, width, format)
    ("Grade",           "grade",              8,   None),
    ("Score",           "score",              8,   None),
    ("County",          "county_name",        14,  None),
    ("State",           "state",              7,   None),
    ("Address",         "address",            35,  None),
    ("Owner Name",      "owner_name",         24,  None),
    ("Case Number",     "case_number",        18,  None),
    ("Sale Date",       "sale_date",          12,  None),
    ("Sale Price",      "final_sale_price",   14,  '"$"#,##0'),
    ("Opening Bid",     "opening_bid",        14,  '"$"#,##0'),
    ("Gross Surplus",   "gross_surplus",      14,  '"$"#,##0'),
    ("Sec. Liens",      "total_secondary_liens", 12, '"$"#,##0'),
    ("Net Surplus",     "net_surplus",        14,  '"$"#,##0'),
    ("Lead Type",       "lead_type",          16,  None),
    ("Claim Status",    "claim_status",       13,  None),
    ("Partial Claim",   "partial_claim",      12,  None),
    ("Doc Status",      "doc_status",         12,  None),
    ("Lien Flags",      "lien_flags_str",     30,  None),
    ("Winner",          "winner_name",        28,  None),
    ("Phone",           "phone1",             15,  None),
    ("Email",           "email",              28,  None),
    ("Mailing Address", "mailing_address",    30,  None),
    ("Next Check",      "next_check",         12,  None),
    ("Outreach Ready",  "outreach_ready",     14,  None),
    ("Parcel ID",       "parcel_id",          18,  None),
    ("Est. Value",      "estimated_value",    14,  '"$"#,##0'),
    ("Prop Type",       "property_type",      16,  None),
    ("PR Enriched",     "pr_enriched",        12,  None),
]


def build_excel(leads: list, date_str: str = None) -> Path:
    """Build the formatted Excel workbook from leads."""
    if not date_str:
        date_str = date.today().isoformat()

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Leads ────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Leads"
    _write_sheet(ws_all, leads, "All Leads")

    # ── Sheet 2: A+ Priority ─────────────────────────────────────────
    priority = [l for l in leads if l.get("grade") in ("A+", "A") and l.get("outreach_ready")]
    ws_pri = wb.create_sheet("A+ Priority")
    _write_sheet(ws_pri, priority, "A+ Priority — Outreach Ready")

    # ── Sheet 3: Florida ─────────────────────────────────────────────
    fl_leads = [l for l in leads if l.get("state") == "FL"]
    ws_fl = wb.create_sheet("Florida")
    _write_sheet(ws_fl, fl_leads, "Florida Leads")

    # ── Sheet 4: Ohio ────────────────────────────────────────────────
    oh_leads = [l for l in leads if l.get("state") == "OH"]
    ws_oh = wb.create_sheet("Ohio")
    _write_sheet(ws_oh, oh_leads, "Ohio Leads")

    # ── Sheet 5: Summary ─────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    _write_summary(ws_sum, leads, date_str)

    # Save
    filename = f"SurplusIQ_Leads_{date_str}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    print(f"✅ Excel export: {filepath}")
    print(f"   {len(leads)} total leads | {len(priority)} A+ priority")
    return filepath


def _write_sheet(ws, leads: list, title: str):
    """Write a leads sheet with formatting."""
    # ── Title Row ─────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell = ws["A1"]
    title_cell.value      = f"SurplusIQ — {title}  |  Generated {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
    title_cell.font       = Font(bold=True, size=12, color=CLR_WHITE)
    title_cell.fill       = PatternFill("solid", fgColor=CLR_HEADER_BG)
    title_cell.alignment  = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Header Row ────────────────────────────────────────────────────
    for col_idx, (header, _, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font       = Font(bold=True, size=10, color=CLR_WHITE)
        cell.fill       = header_fill()
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 28

    # Freeze panes (keep header visible)
    ws.freeze_panes = "A3"

    # ── Data Rows ─────────────────────────────────────────────────────
    for row_idx, lead in enumerate(leads, start=3):
        grade       = lead.get("grade", "C")
        claim_status = lead.get("claim_status", "none")
        is_alt       = (row_idx % 2 == 0)
        base_fill    = PatternFill("solid", fgColor=CLR_ROW_ALT) if is_alt else PatternFill("solid", fgColor=CLR_WHITE)

        for col_idx, (_, field_key, _, fmt) in enumerate(COLUMNS, start=1):
            value = lead.get(field_key, "")

            # Format booleans
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            elif value is None:
                value = ""
            elif isinstance(value, list):
                value = " | ".join(str(v) for v in value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.font      = Font(size=9)

            # Apply number format
            if fmt:
                cell.number_format = fmt

            # Grade column coloring
            if field_key == "grade":
                cell.fill      = grade_fill(grade)
                cell.font      = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif field_key == "score":
                cell.fill      = base_fill
                cell.font      = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif field_key == "claim_status":
                cell.fill      = claim_fill(claim_status)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif field_key == "net_surplus":
                cell.fill = base_fill
                cell.font = Font(bold=True, size=9, color="1A6B3C")
            elif field_key == "outreach_ready":
                if value == "Yes":
                    cell.fill = PatternFill("solid", fgColor="D4EFDF")
                    cell.font = Font(bold=True, size=9, color="1A6B3C")
                else:
                    cell.fill = base_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif field_key in ("final_sale_price", "opening_bid", "gross_surplus", "total_secondary_liens", "estimated_value"):
                cell.fill = base_fill
            else:
                cell.fill = base_fill

        ws.row_dimensions[row_idx].height = 18

    # Auto-filter on header row
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{len(leads)+2}"


def _write_summary(ws, leads: list, date_str: str):
    """Write a summary statistics sheet."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    def s(row, col, val, bold=False, fmt=None, bg=None, color=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font      = Font(bold=bold, size=10, color=color or "1A1A1A")
        cell.alignment = Alignment(vertical="center")
        cell.border    = thin_border()
        if fmt:
            cell.number_format = fmt
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        return cell

    # Title
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value      = f"SurplusIQ — Summary Report | {date_str}"
    t.font       = Font(bold=True, size=13, color=CLR_WHITE)
    t.fill       = PatternFill("solid", fgColor=CLR_HEADER_BG)
    t.alignment  = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    row = 3

    # Overall stats
    total_surplus = sum(l.get("net_surplus", 0) or 0 for l in leads)
    a_plus        = sum(1 for l in leads if l.get("grade") == "A+")
    a             = sum(1 for l in leads if l.get("grade") == "A")
    b             = sum(1 for l in leads if l.get("grade") == "B")
    c             = sum(1 for l in leads if l.get("grade") == "C")
    outreach_rdy  = sum(1 for l in leads if l.get("outreach_ready"))
    no_claim      = sum(1 for l in leads if l.get("claim_status") == "none")
    partial       = sum(1 for l in leads if l.get("claim_status") == "partial")
    docs_ready    = sum(1 for l in leads if l.get("doc_available"))

    for label, val, fmt, bg in [
        ("Generated",           datetime.now().strftime("%B %d, %Y"), None, None),
        ("Total Leads",         len(leads),       None,         None),
        ("Total Net Surplus",   total_surplus,    '"$"#,##0',   "D4EFDF"),
        ("A+ Leads",            a_plus,           None,         "D4EFDF"),
        ("A Leads",             a,                None,         "EBF5FB"),
        ("B Leads",             b,                None,         "FEF9E7"),
        ("C Leads",             c,                None,         "FDEDEC"),
        ("Outreach Ready",      outreach_rdy,     None,         "D4EFDF"),
        ("No Claim Filed",      no_claim,         None,         "D4EFDF"),
        ("Partial Claims",      partial,          None,         "FEF9E7"),
        ("Docs Retrieved",      docs_ready,       None,         None),
    ]:
        s(row, 1, label, bold=True)
        s(row, 2, val,   fmt=fmt, bg=bg)
        row += 1

    row += 1

    # Per-county breakdown
    s(row, 1, "County Breakdown", bold=True, bg=CLR_HEADER_BG, color=CLR_WHITE)
    s(row, 2, "Leads",            bold=True, bg=CLR_HEADER_BG, color=CLR_WHITE)
    s(row, 3, "Net Surplus",      bold=True, bg=CLR_HEADER_BG, color=CLR_WHITE)
    s(row, 4, "A+ Count",         bold=True, bg=CLR_HEADER_BG, color=CLR_WHITE)
    row += 1

    counties = {}
    for lead in leads:
        cname = f"{lead.get('county_name', '?')} ({lead.get('state', '?')})"
        if cname not in counties:
            counties[cname] = {"leads": 0, "surplus": 0, "aplus": 0}
        counties[cname]["leads"]  += 1
        counties[cname]["surplus"] += lead.get("net_surplus", 0) or 0
        if lead.get("grade") == "A+":
            counties[cname]["aplus"] += 1

    for cname, stats in sorted(counties.items(), key=lambda x: x[1]["surplus"], reverse=True):
        s(row, 1, cname)
        s(row, 2, stats["leads"])
        s(row, 3, stats["surplus"], fmt='"$"#,##0')
        s(row, 4, stats["aplus"])
        row += 1

    ws.freeze_panes = "A3"


if __name__ == "__main__":
    # Test with fake data
    test_leads = [
        {
            "grade": "A+", "score": 92, "county_name": "Miami-Dade", "state": "FL",
            "address": "4821 NW 7th Ave, Miami, FL", "owner_name": "Marcus T. Johnson",
            "case_number": "2024-CA-048821", "sale_date": "2026-04-03",
            "final_sale_price": 312000, "opening_bid": 263800,
            "gross_surplus": 48200, "total_secondary_liens": 0, "net_surplus": 48200,
            "lead_type": "Mortgage Foreclosure", "claim_status": "none",
            "partial_claim": False, "doc_status": "retrieved", "doc_available": True,
            "lien_flags_str": "None", "winner_name": "MDH Real Estate Partners LLC",
            "phone1": "(305) 555-0142", "email": "", "mailing_address": "4821 NW 7th Ave",
            "next_check": "complete", "outreach_ready": True,
            "parcel_id": "01-3127-001-0010", "estimated_value": 285000,
            "property_type": "SFR", "pr_enriched": True,
        },
        {
            "grade": "A", "score": 74, "county_name": "Cuyahoga", "state": "OH",
            "address": "1203 Riverside Dr, Cleveland, OH", "owner_name": "Patricia L. Chen",
            "case_number": "CV-2024-001234", "sale_date": "2026-04-01",
            "final_sale_price": 198000, "opening_bid": 166500,
            "gross_surplus": 31500, "total_secondary_liens": 0, "net_surplus": 31500,
            "lead_type": "Mortgage Foreclosure", "claim_status": "none",
            "partial_claim": False, "doc_status": "retrieved", "doc_available": True,
            "lien_flags_str": "None", "winner_name": "Cleveland Realty Group LLC",
            "phone1": "(216) 555-0188", "email": "", "mailing_address": "1203 Riverside Dr",
            "next_check": "complete", "outreach_ready": True,
            "parcel_id": "120-12-001", "estimated_value": 175000,
            "property_type": "SFR", "pr_enriched": True,
        },
    ]

    filepath = build_excel(test_leads, "test")
    print(f"Test Excel built: {filepath}")
